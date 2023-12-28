from PySide6.QtWidgets import QWidget, QMessageBox
from PySide6.QtCore import QThread, Signal, QDate
from salesUpdate_ui import Ui_Form
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import create_engine
from sqlalchemy.engine import URL
import pandas as pd
import json, datetime


class Worker(QThread):
    task = Signal(str)
    progress = Signal(int)
    
    def __init__(self, root_path, dateFrom:QDate, dateTo:QDate):
        super().__init__()
        self._root_path = root_path
        self._dateFrom = dateFrom
        self._dateTo = dateTo

    def run(self):
        # Create RMN INV table
        self.task.emit('Load database auth info...')
        self.progress.emit(0)

        with open(self._root_path+'appdata/db_auth.json') as f:
            temp = json.load(f)
            server = temp['server']
            database = temp['database'] 
            username = temp['username'] 
            password = temp['password']

        # Create connection to database
        self.task.emit('Create connection...')
        self.progress.emit(10)

        connection_string = 'DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password
        connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": connection_string})

        engine = create_engine(connection_url)

        # Query POS data
        self.task.emit('Get POS data...')
        self.progress.emit(20)

        query = f"""
        SELECT 
            Item.ItemLookupCode AS 'Item Lookup Code',
            Item.Price AS 'Price',
            Item.Quantity AS 'Qty On Hand',
            Item.SubDescription3 AS 'Display',
            Item.SubDescription2 AS 'Comp Inv {datetime.date.today().strftime("%m%d")}',
            Item.Description AS 'Description',
            Item.ExtendedDescription AS 'Extended Description',
            Item.BinLocation AS 'Bin Location',
            sl.ReorderNumber AS 'Reorder Number',
            Item.SubDescription1 AS 'BRAND',
            dp.Name AS 'Departments',
            sp.Code AS 'Supplier Code',
            sp.SupplierName AS 'Supplier Name'
        FROM
            dbo.Item Item
            LEFT JOIN dbo.SupplierList sl ON Item.ID=sl.ItemID AND Item.SupplierID=sl.SupplierID
            LEFT JOIN dbo.Department dp ON Item.DepartmentID=dp.ID
            LEFT JOIN dbo.Supplier sp ON Item.SupplierID=sp.ID
        WHERE
            Item.DepartmentID IN (2, 4, 6) 
            AND Item.Inactive = 0
        ORDER BY
            ItemLookupCode;
        """

        with engine.connect() as conn, conn.begin():  
            fromPOS = pd.read_sql(query, conn, dtype={'Qty On Hand':'int64'},)

        fromPOS['Display'] = fromPOS['Display'].str.strip()
        fromPOS.loc[fromPOS['Display'].str.startswith('0'), 'Display'] = '0'
        fromPOS.loc[fromPOS['Display'].str.startswith('1'), 'Display'] = '1'
        fromPOS.loc[fromPOS['Display']=='', 'Display']='0'

        fromPOS['ITEM QTY'] = fromPOS['Qty On Hand']-fromPOS['Display'].astype('int64')
        fromPOS.loc[fromPOS['ITEM QTY']<0, 'ITEM QTY'] = 0

        temp = fromPOS[['Reorder Number', 'ITEM QTY']].groupby('Reorder Number').sum()
        temp.columns=['FIN TOT QTY']

        rmh_inv = fromPOS.merge(temp, on='Reorder Number', how='left')
        rmh_inv['FIN TOT QTY'] = rmh_inv['FIN TOT QTY'].fillna(0).astype('int64')

        # Query sales data
        self.task.emit('Get Sales data...')
        self.progress.emit(40)
        query = f"""
        SELECT 
            FORMAT(hs.DateTransferred, 'yyyy-MM-dd') AS 'Date',
            hs.ItemLookupCode AS 'Item Lookup Code',
            hs.ItemDescription AS 'Description',
            hs.Quantity AS 'QTY SOLD',
            hs.DepartmentName AS 'Department',
            FORMAT(hs.DateTransferred, 'yyMM') AS 'yymm'
        FROM 
            dbo.ViewItemMovementHistory hs
        WHERE
            hs.DepartmentName IN ('Braids', 'Hair Extensions', 'Wigs')
            AND hs.Type=99
            AND hs.DateTransferred BETWEEN '{self._dateFrom.year()}-{self._dateFrom.month()}-{self._dateFrom.day()}' AND '{self._dateTo.year()}-{self._dateTo.month()}-{self._dateTo.day()} 23:59:59'
        ORDER BY
            hs.DateTransferred;
        """

        with engine.connect() as conn, conn.begin():  
            item_history = pd.read_sql(query, conn, dtype={'QTY SOLD': 'int64'})

        merged = item_history.merge(rmh_inv[['Item Lookup Code', 'Reorder Number', 'Supplier Name', 'Qty On Hand',\
                                    f'Comp Inv {datetime.date.today().strftime("%m%d")}', 'FIN TOT QTY']],
                                    how='inner', on='Item Lookup Code', sort='Date')
        
        merged.rename({'Reorder Number':'Item',
               'Supplier Name':'Company',
               'Qty On Hand':'RMH_Inv',
               f'Comp Inv {datetime.date.today().strftime("%m%d")}':'Comp Inv',
               'FIN TOT QTY':'Item Tot'}, axis='columns', inplace=True)
        
        merged.insert(7, 'Color', pd.NA)

        for i in range(len(merged)):
            merged.loc[i,'Color'] = merged.loc[i, 'Description'][len(merged['Item'][i])+1:]

        merged['Item_Inv'] = merged['Item']+'('+merged['Item Tot'].astype(str)+')'
        merged['Color_Inv'] = merged['Color']+'('+merged['RMH_Inv'].astype(str)+' - '+merged['Comp Inv'].astype(str)+') - '+merged['Item Lookup Code'].astype(str)
        merged['st_itm_cmp'] = '(' + merged['RMH_Inv'].astype(str) + ') ' + merged['Item'].astype(str) + ' ('+merged['Comp Inv'].astype(str) + ')'

        merged.sort_values(by="Date", ignore_index=True,inplace=True)

        
        # Create STORE Sales report
        self.task.emit('Creating Report...')
        self.progress.emit(50)

        wb = load_workbook(self._root_path+'appdata/STORE sales template.xlsx')

        ws1=wb.worksheets[1]
        ws2=wb.worksheets[2]

        ws1.delete_rows(2,ws1.max_row-1)
        ws2.delete_rows(2, ws2.max_row-1)
        self.progress.emit(60)

        for r in dataframe_to_rows(merged, index=False, header=False):
            ws1.append(r)
        self.progress.emit(70)

        for r in dataframe_to_rows(rmh_inv, index=False, header=False):
            ws2.append(r)

        self.progress.emit(80)

        wb.save(self._root_path+f'STORE Sales_{datetime.datetime.today().strftime("%m%d%y")}.xlsx')
        self.task.emit(f'Sales Report Created as "STORE Sales_{datetime.datetime.today().strftime("%m%d%y")}.xlsx"')
        self.progress.emit(100)
        

class SalesUpdateWindow(QWidget):
    def __init__(self, root_path):
        super().__init__()
        self._root_path = root_path
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        self.ui.dateEdit.setDate(QDate(2021, 1,1))
        self.ui.dateEdit_2.setDate(QDate.currentDate())

        self.ui.createButton.clicked.connect(self.create_report)
        

    def create_report(self):
        self.ui.createButton.setEnabled(False)
        self.ui.closeButton.setEnabled(False)
        self.ui.dateEdit.setEnabled(False)
        self.ui.dateEdit_2.setEnabled(False)

        self.worker = Worker(self._root_path, self.ui.dateEdit.date(), self.ui.dateEdit_2.date())
        self.worker.task.connect(self.ui.label_3.setText)
        self.worker.progress.connect(self.ui.progressBar.setValue)
        self.worker.finished.connect(self.report_created)
        self.worker.start()

    def report_created(self):
        self.ui.createButton.setEnabled(True)
        self.ui.closeButton.setEnabled(True)
        self.ui.dateEdit.setEnabled(True)
        self.ui.dateEdit_2.setEnabled(True)

        if self.ui.progressBar.value() == self.ui.progressBar.maximum():
            QMessageBox.information(self, "Info", "Complete")

        del self.worker